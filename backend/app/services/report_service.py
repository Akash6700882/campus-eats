import csv
import io
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.repositories.analytics_repository import AnalyticsRepository
from app.schemas.analytics import BestSellingFood, DailyRevenue, OrderStatusCount
from app.schemas.report import ReportPeriod, ReportResponse

# Trailing windows (not calendar day/week/month/year) — "monthly" is the
# trailing 30 days rather than the current calendar month, "yearly" the
# trailing 365 days, etc. Simpler to reason about and consistent with how
# AnalyticsRepository.revenue_by_day already defines its trailing-7-day
# dashboard window.
_PERIOD_DAYS = {
    ReportPeriod.DAILY: 1,
    ReportPeriod.WEEKLY: 7,
    ReportPeriod.MONTHLY: 30,
    ReportPeriod.YEARLY: 365,
}


class ReportService:
    def __init__(self, analytics_repo: AnalyticsRepository):
        self.analytics_repo = analytics_repo

    @staticmethod
    def _resolve_range(period: ReportPeriod) -> tuple[datetime, datetime]:
        end = datetime.now(timezone.utc)
        start = end - timedelta(days=_PERIOD_DAYS[period])
        return start, end

    async def build_report(self, period: ReportPeriod) -> ReportResponse:
        start, end = self._resolve_range(period)

        status_counts = await self.analytics_repo.order_counts_by_status_in_range(start, end)
        total_revenue = await self.analytics_repo.total_revenue_in_range(start, end)
        daily = await self.analytics_repo.revenue_by_day_in_range(start, end)
        best_selling = await self.analytics_repo.best_selling_foods_in_range(start, end)

        return ReportResponse(
            period=period,
            start_date=start.date().isoformat(),
            end_date=end.date().isoformat(),
            total_orders=sum(status_counts.values()),
            total_revenue=total_revenue,
            orders_by_status=[OrderStatusCount(status=s, count=c) for s, c in status_counts.items()],
            best_selling_foods=[BestSellingFood(name=n, quantity_sold=q, revenue=r) for n, q, r in best_selling],
            daily_breakdown=[DailyRevenue(date=d, revenue=r, order_count=c) for d, r, c in daily],
        )

    @staticmethod
    def render_csv(report: ReportResponse) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Campus Eats Report", report.period.value, report.start_date, report.end_date])
        writer.writerow([])
        writer.writerow(["Total Orders", report.total_orders])
        writer.writerow(["Total Revenue", report.total_revenue])
        writer.writerow([])
        writer.writerow(["Orders by status"])
        writer.writerow(["Status", "Count"])
        for row in report.orders_by_status:
            writer.writerow([row.status, row.count])
        writer.writerow([])
        writer.writerow(["Best-selling foods"])
        writer.writerow(["Food", "Quantity sold", "Revenue"])
        for row in report.best_selling_foods:
            writer.writerow([row.name, row.quantity_sold, row.revenue])
        writer.writerow([])
        writer.writerow(["Daily breakdown"])
        writer.writerow(["Date", "Revenue", "Order count"])
        for row in report.daily_breakdown:
            writer.writerow([row.date, row.revenue, row.order_count])
        return buffer.getvalue().encode("utf-8")

    @staticmethod
    def render_excel(report: ReportResponse) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        bold = Font(bold=True)

        ws.append([f"Campus Eats Report — {report.period.value}", report.start_date, report.end_date])
        ws.append([])
        ws.append(["Total Orders", report.total_orders])
        ws.append(["Total Revenue", report.total_revenue])
        ws.append([])
        ws.append(["Orders by status"])
        ws.cell(ws.max_row, 1).font = bold
        ws.append(["Status", "Count"])
        for row in report.orders_by_status:
            ws.append([row.status, row.count])
        ws.append([])
        ws.append(["Best-selling foods"])
        ws.cell(ws.max_row, 1).font = bold
        ws.append(["Food", "Quantity sold", "Revenue"])
        for row in report.best_selling_foods:
            ws.append([row.name, row.quantity_sold, row.revenue])
        ws.append([])
        ws.append(["Daily breakdown"])
        ws.cell(ws.max_row, 1).font = bold
        ws.append(["Date", "Revenue", "Order count"])
        for row in report.daily_breakdown:
            ws.append([row.date, row.revenue, row.order_count])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def render_pdf(report: ReportResponse) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f"Campus Eats — {report.period.value.title()} Report", styles["Title"]),
            Paragraph(f"{report.start_date} to {report.end_date}", styles["Normal"]),
            Spacer(1, 12),
            Paragraph(f"Total Orders: {report.total_orders}", styles["Normal"]),
            Paragraph(f"Total Revenue: Rs. {report.total_revenue:.2f}", styles["Normal"]),
            Spacer(1, 12),
        ]

        def section(title: str, header: list[str], rows: list[list[str]]) -> None:
            elements.append(Paragraph(title, styles["Heading3"]))
            table = Table([header, *rows])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f97316")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 12))

        section(
            "Orders by status",
            ["Status", "Count"],
            [[row.status, str(row.count)] for row in report.orders_by_status] or [["—", "—"]],
        )
        section(
            "Best-selling foods",
            ["Food", "Quantity sold", "Revenue"],
            [[row.name, str(row.quantity_sold), f"Rs. {row.revenue:.2f}"] for row in report.best_selling_foods]
            or [["—", "—", "—"]],
        )
        section(
            "Daily breakdown",
            ["Date", "Revenue", "Order count"],
            [[row.date, f"Rs. {row.revenue:.2f}", str(row.order_count)] for row in report.daily_breakdown]
            or [["—", "—", "—"]],
        )

        doc.build(elements)
        return buffer.getvalue()
